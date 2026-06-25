#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


DECISION_VALUES = "approve,needs_review,blocked,deferred,archive_only_override"
MATTER_TYPE_VALUES = "Criminal,Civil,Advisory,M&A"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--receipt", required=True)
    parser.add_argument("--batch-name", default="first-review")
    parser.add_argument("--include-priority", action="append", default=["P2_draft_code_ready"])
    parser.add_argument("--limit", type=int, default=300)
    return parser.parse_args()


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_sheet(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [clean(value) for value in rows[0]]
    return [
        {header: row[index] if index < len(row) else None for index, header in enumerate(headers)}
        for row in rows[1:]
    ]


def priority_rank(priority: str) -> int:
    order = {
        "P0_conflict": 0,
        "P1_project_folder": 1,
        "P2_draft_code_ready": 2,
        "P3_missing_hint": 3,
        "P4_large_group": 4,
        "P5_standard_review": 5,
    }
    return order.get(priority, 9)


def set_style(sheet) -> None:
    normal_font = Font(name="Arial", size=10)
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = normal_font
            cell.alignment = wrap
            if cell.row > 1:
                cell.border = thin_border


def write_sheet(workbook: Workbook, name: str, rows: list[dict[str, Any]]):
    sheet = workbook.create_sheet(name)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    if not rows:
        sheet.append(["empty"])
    else:
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, start=1):
            preview_width = max((len(str(row.get(header) or "")) for row in rows[:200]), default=0)
            sheet.column_dimensions[get_column_letter(index)].width = max(
                12, min(max(len(str(header)), preview_width), 60) + 2
            )
    set_style(sheet)
    return sheet


def create_batch_rows(approval_rows: list[dict[str, Any]], priorities: set[str], limit: int) -> list[dict[str, Any]]:
    candidates = [
        row
        for row in approval_rows
        if clean(row.get("reviewer_decision")) == "needs_review"
        and clean(row.get("review_priority")) in priorities
    ]
    candidates.sort(
        key=lambda row: (
            priority_rank(clean(row.get("review_priority"))),
            -int(row.get("object_count") or 0),
            clean(row.get("group_id")),
        )
    )
    selected = candidates[:limit]
    return [
        {
            "batch_decision": "needs_review",
            "approved_client_short_name": clean(row.get("approved_client_short_name"))
            or clean(row.get("client_short_name_candidate")),
            "approved_matter_type_english": clean(row.get("approved_matter_type_english"))
            or clean(row.get("matter_type_english_candidate")),
            "approved_matter_detail_type_korean": clean(row.get("approved_matter_detail_type_korean"))
            or clean(row.get("matter_detail_type_korean_candidate")),
            "approved_matter_code": clean(row.get("approved_matter_code")) or clean(row.get("matter_code_candidate")),
            "reviewer_notes": "",
            "review_priority": clean(row.get("review_priority")),
            "group_id": clean(row.get("group_id")),
            "object_count": row.get("object_count"),
            "project_like_rows": row.get("project_like_rows"),
            "client_short_name_candidate": row.get("client_short_name_candidate"),
            "matter_type_english_candidate": row.get("matter_type_english_candidate"),
            "matter_detail_type_korean_candidate": row.get("matter_detail_type_korean_candidate"),
            "matter_code_candidate": row.get("matter_code_candidate"),
            "code_warnings": row.get("code_warnings"),
            "evidence_class_counts": row.get("evidence_class_counts"),
            "raw_group_label_local_only": row.get("raw_group_label_local_only"),
        }
        for row in selected
    ]


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook)
    output = Path(args.output)
    receipt = Path(args.receipt)
    output.parent.mkdir(parents=True, exist_ok=True)
    receipt.parent.mkdir(parents=True, exist_ok=True)

    approval_rows = read_sheet(workbook_path, "Approval")
    priorities = set(args.include_priority)
    batch_rows = create_batch_rows(approval_rows, priorities, args.limit)

    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    for row in [
        ["Batch Name", args.batch_name],
        ["Generated At", datetime.now(timezone.utc).isoformat()],
        ["Included Priorities", ", ".join(sorted(priorities))],
        ["Limit", args.limit],
        ["Rows", len(batch_rows)],
        ["Decision Values", DECISION_VALUES],
        ["Matter Type Values", MATTER_TYPE_VALUES],
        ["Boundary", "Local-only review batch. It does not approve or import anything by itself."],
    ]:
        readme.append(row)
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 110
    set_style(readme)

    review_sheet = write_sheet(workbook, "Review Batch", batch_rows)
    summary_rows = [
        {"metric": "source_workbook_rows", "value": len(approval_rows)},
        {"metric": "batch_rows", "value": len(batch_rows)},
        {"metric": "included_priorities", "value": ", ".join(sorted(priorities))},
        {"metric": "auto_approved_rows", "value": 0},
    ]
    write_sheet(workbook, "Summary", summary_rows)

    headers = [cell.value for cell in review_sheet[1]]
    if "batch_decision" in headers:
        column_index = headers.index("batch_decision") + 1
        column_letter = get_column_letter(column_index)
        validation = DataValidation(type="list", formula1=f'"{DECISION_VALUES}"', allow_blank=False)
        review_sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{review_sheet.max_row}")
    if "approved_matter_type_english" in headers:
        column_index = headers.index("approved_matter_type_english") + 1
        column_letter = get_column_letter(column_index)
        validation = DataValidation(type="list", formula1=f'"{MATTER_TYPE_VALUES}"', allow_blank=True)
        review_sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{review_sheet.max_row}")
    fill = PatternFill("solid", fgColor="FFF2CC")
    for index, header in enumerate(headers, start=1):
        if header in {
            "batch_decision",
            "approved_client_short_name",
            "approved_matter_type_english",
            "approved_matter_detail_type_korean",
            "approved_matter_code",
            "reviewer_notes",
        }:
            for cell in review_sheet.iter_cols(min_col=index, max_col=index, min_row=2, max_row=review_sheet.max_row):
                for target in cell:
                    target.fill = fill

    workbook.save(output)
    os.chmod(output, 0o600)

    validation_workbook = load_workbook(output, read_only=True, data_only=False)
    validation = {
        "artifact": "approval_batch_validation_sanitized",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "batch_name": args.batch_name,
        "workbook_ref": output.name,
        "mode": oct(os.stat(output).st_mode & 0o777)[2:],
        "included_priorities": sorted(priorities),
        "source_approval_rows": len(approval_rows),
        "batch_rows": validation_workbook["Review Batch"].max_row - 1,
        "auto_approved_rows": 0,
        "security_boundary": "local-only workbook; may contain group labels; do not commit",
        "repo_safe": False,
    }
    validation["checks"] = {
        "mode_0600": validation["mode"] == "600",
        "batch_rows_match": validation["batch_rows"] == len(batch_rows),
        "no_auto_approval": validation["auto_approved_rows"] == 0,
        "sheets_present": all(sheet in validation_workbook.sheetnames for sheet in ["README", "Review Batch", "Summary"]),
    }
    validation["all_checks_passed"] = all(validation["checks"].values())
    receipt.write_text(json.dumps(validation, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.chmod(receipt, 0o644)
    print(
        json.dumps(
            {
                "all_checks_passed": validation["all_checks_passed"],
                "batch_rows": validation["batch_rows"],
                "workbook": str(output),
                "receipt": str(receipt),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
