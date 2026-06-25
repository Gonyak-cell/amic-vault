#!/usr/bin/env python3
from __future__ import annotations

import argparse
import gzip
import hashlib
import json
import os
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DECISIONS = {"approve", "needs_review", "blocked", "deferred", "archive_only_override"}
ALLOWED_MATTER_TYPES = {"Criminal", "Civil", "Advisory", "M&A"}
REQUIRED_APPROVAL_COLUMNS = [
    "reviewer_decision",
    "approved_client_short_name",
    "approved_matter_type_english",
    "approved_matter_detail_type_korean",
    "approved_matter_code",
    "group_id",
    "object_count",
    "matter_code_candidate",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--local-dir", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--receipt", required=True)
    return parser.parse_args()


def read_ndjson_gz(path: Path) -> list[dict[str, Any]]:
    with gzip.open(path, "rt", encoding="utf-8") as handle:
        return [json.loads(line) for line in handle if line.strip()]


def write_ndjson_gz(path: Path, rows: list[dict[str, Any]]) -> None:
    payload = "\n".join(json.dumps(row, ensure_ascii=False, sort_keys=True) for row in rows)
    if payload:
        payload += "\n"
    with gzip.open(path, "wt", encoding="utf-8") as handle:
        handle.write(payload)
    os.chmod(path, 0o600)


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sha256_hex(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def load_approval_rows(workbook_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    if "Approval" not in workbook.sheetnames:
        return [], ["missing_approval_sheet"]
    sheet = workbook["Approval"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], ["empty_approval_sheet"]
    headers = [clean(value) for value in rows[0]]
    missing = [column for column in REQUIRED_APPROVAL_COLUMNS if column not in headers]
    if missing:
        return [], [f"missing_column:{column}" for column in missing]
    parsed: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[1:], start=2):
        item = {header: row[index] if index < len(row) else None for index, header in enumerate(headers)}
        item["_row_index"] = row_index
        parsed.append(item)
    return parsed, []


def approval_values(row: dict[str, Any]) -> dict[str, str]:
    client = clean(row.get("approved_client_short_name")) or clean(row.get("client_short_name_candidate"))
    matter_type = clean(row.get("approved_matter_type_english")) or clean(row.get("matter_type_english_candidate"))
    detail = clean(row.get("approved_matter_detail_type_korean")) or clean(
        row.get("matter_detail_type_korean_candidate")
    )
    code = clean(row.get("approved_matter_code")) or clean(row.get("matter_code_candidate"))
    return {
        "client_short_name": client,
        "matter_type_english": matter_type,
        "matter_detail_type_korean": detail,
        "matter_code": code,
    }


def validate_approved_row(row: dict[str, Any]) -> list[str]:
    values = approval_values(row)
    blockers = []
    if not values["client_short_name"]:
        blockers.append("missing_client_short_name")
    if not values["matter_type_english"]:
        blockers.append("missing_matter_type_english")
    elif values["matter_type_english"] not in ALLOWED_MATTER_TYPES:
        blockers.append("unsupported_matter_type_english")
    if not values["matter_detail_type_korean"]:
        blockers.append("missing_matter_detail_type_korean")
    if not values["matter_code"]:
        blockers.append("missing_matter_code")
    parts = values["matter_code"].split("/") if values["matter_code"] else []
    if values["matter_code"] and (len(parts) != 3 or any(not part.strip() for part in parts)):
        blockers.append("invalid_matter_code_format")
    if values["matter_code"] and len(values["matter_code"]) > 120:
        blockers.append("matter_code_over_120_chars")
    expected = (
        f"{values['client_short_name']}/{values['matter_type_english']}/{values['matter_detail_type_korean']}"
        if values["client_short_name"] and values["matter_type_english"] and values["matter_detail_type_korean"]
        else ""
    )
    if values["matter_code"] and expected and values["matter_code"] != expected:
        blockers.append("matter_code_does_not_match_approved_segments")
    return blockers


def source_row_matches_group(row: dict[str, Any], approved_groups: set[str]) -> bool:
    groups = row.get("groups") or {}
    return any(group_id in approved_groups for group_id in groups.values())


def deepest_approved_group(row: dict[str, Any], approved_groups: set[str]) -> str | None:
    groups = row.get("groups") or {}
    depth_group_pairs = []
    for key, group_id in groups.items():
        if group_id not in approved_groups:
            continue
        try:
            depth = int(str(key).split("_", 1)[1])
        except (IndexError, ValueError):
            depth = 0
        depth_group_pairs.append((depth, group_id))
    if not depth_group_pairs:
        return None
    return max(depth_group_pairs, key=lambda item: item[0])[1]


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook)
    local_dir = Path(args.local_dir)
    output_dir = Path(args.output_dir)
    receipt_path = Path(args.receipt)
    output_dir.mkdir(parents=True, exist_ok=True)
    receipt_path.parent.mkdir(parents=True, exist_ok=True)

    approval_rows, workbook_blockers = load_approval_rows(workbook_path)
    decision_counts = Counter()
    invalid_decision_rows = []
    approved_groups: dict[str, dict[str, str]] = {}
    approved_blockers: dict[str, list[str]] = {}

    for row in approval_rows:
        decision = clean(row.get("reviewer_decision")) or "needs_review"
        decision_counts[decision] += 1
        if decision not in DECISIONS:
            invalid_decision_rows.append(row["_row_index"])
            continue
        if decision != "approve":
            continue
        group_id = clean(row.get("group_id"))
        blockers = validate_approved_row(row)
        if blockers:
            approved_blockers[group_id or f"row:{row['_row_index']}"] = blockers
            continue
        approved_groups[group_id] = approval_values(row)

    matter_code_counts = Counter(values["matter_code"] for values in approved_groups.values())
    duplicate_codes = [code for code, count in matter_code_counts.items() if count > 1]
    duplicate_group_rows = sum(matter_code_counts[code] for code in duplicate_codes)

    approved_groups_without_blockers = {
        group_id: values for group_id, values in approved_groups.items() if group_id not in approved_blockers
    }
    target_matter_codes: dict[str, dict[str, Any]] = {}
    for group_id, values in sorted(approved_groups_without_blockers.items()):
        matter_code = values["matter_code"]
        target = target_matter_codes.setdefault(
            matter_code,
            {
                "matter_code_hash": sha256_hex(matter_code),
                "client_short_name": values["client_short_name"],
                "matter_type_english": values["matter_type_english"],
                "matter_detail_type_korean": values["matter_detail_type_korean"],
                "target_resolution_state": "pending_db_target_resolution",
                "approved_group_ids": [],
                "approved_group_count": 0,
            },
        )
        target["approved_group_ids"].append(group_id)
        target["approved_group_count"] += 1

    row_lanes = read_ndjson_gz(local_dir / "bulk-row-lanes.local.ndjson.gz")
    approved_scope_rows = []
    for row in row_lanes:
        if row.get("lane") != "mapping_required":
            continue
        matched_group = deepest_approved_group(row, set(approved_groups_without_blockers))
        if not matched_group:
            continue
        values = approved_groups_without_blockers[matched_group]
        approved_scope_rows.append(
            {
                "source_object_hash": row.get("source_object_hash"),
                "group_id": matched_group,
                "extension": row.get("extension"),
                "size_bytes": row.get("size_bytes"),
                "client_short_name": values["client_short_name"],
                "matter_type_english": values["matter_type_english"],
                "matter_detail_type_korean": values["matter_detail_type_korean"],
                "matter_code": values["matter_code"],
                "raw": row.get("raw"),
            }
        )

    approved_groups_path = output_dir / "approved-groups.local.ndjson.gz"
    approved_scope_path = output_dir / "approved-import-scope.local.ndjson.gz"
    target_plan_path = output_dir / "approved-target-resolution-plan.local.ndjson.gz"
    write_checklist_path = output_dir / "approved-write-readiness.local.ndjson.gz"

    write_ndjson_gz(
        approved_groups_path,
        [
            {"group_id": group_id, **values, "review_state": "approved"}
            for group_id, values in sorted(approved_groups_without_blockers.items())
        ],
    )
    write_ndjson_gz(approved_scope_path, approved_scope_rows)
    write_ndjson_gz(
        target_plan_path,
        [target_matter_codes[code] for code in sorted(target_matter_codes)],
    )
    write_ndjson_gz(
        write_checklist_path,
        [
            {
                "gate": "approved_groups_present",
                "status": "pass" if approved_groups_without_blockers else "blocked",
                "count": len(approved_groups_without_blockers),
            },
            {
                "gate": "approved_import_scope_present",
                "status": "pass" if approved_scope_rows else "blocked",
                "count": len(approved_scope_rows),
            },
            {
                "gate": "target_resolution_dry_run",
                "status": "pending" if approved_scope_rows else "blocked",
                "count": 0,
            },
        ],
    )

    status = "ready_for_target_resolution" if approved_scope_rows and not approved_blockers else "blocked_pending_approval"
    receipt = {
        "artifact": "onedrive_approval_ingest_sanitized",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "workbook_ref": workbook_path.name,
        "status": status,
        "decision_counts": dict(sorted(decision_counts.items())),
        "approval_rows": len(approval_rows),
        "approved_group_count": len(approved_groups_without_blockers),
        "approved_target_matter_code_count": len(target_matter_codes),
        "approved_source_rows": len(approved_scope_rows),
        "invalid_decision_row_count": len(invalid_decision_rows),
        "approved_blocker_group_count": len(approved_blockers),
        "workbook_blockers": workbook_blockers,
        "duplicate_approved_matter_code_count": len(duplicate_codes),
        "duplicate_approved_matter_code_group_rows": duplicate_group_rows,
        "duplicate_approved_matter_code_policy": "allowed_multiple_groups_resolve_to_one_matter_code",
        "local_outputs": {
            "approved_groups_ref": approved_groups_path.name,
            "approved_import_scope_ref": approved_scope_path.name,
            "approved_target_resolution_plan_ref": target_plan_path.name,
            "approved_write_readiness_ref": write_checklist_path.name,
            "permissions": "0600",
            "commit_allowed": False,
        },
        "not_executed": [
            "Vault DB write",
            "Vault storage write",
            "Matter app write",
            "customer-wide import",
            "source-of-truth cutover",
        ],
        "sanitization": "No raw object keys, raw paths, filenames, document contents, bucket names, account IDs, ARNs, tokens, cookies, or secrets are included.",
    }
    receipt["checks"] = {
        "workbook_parsed": not workbook_blockers,
        "no_invalid_decisions": len(invalid_decision_rows) == 0,
        "no_approved_blockers": len(approved_blockers) == 0,
        "archive_rows_not_promoted": True,
        "local_outputs_mode_0600": all(
            (path.stat().st_mode & 0o777) == 0o600
            for path in [approved_groups_path, approved_scope_path, target_plan_path, write_checklist_path]
        ),
    }
    receipt["all_checks_passed"] = all(receipt["checks"].values())
    receipt_path.write_text(json.dumps(receipt, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.chmod(receipt_path, 0o644)
    print(
        json.dumps(
            {
                "status": receipt["status"],
                "all_checks_passed": receipt["all_checks_passed"],
                "decision_counts": receipt["decision_counts"],
                "approved_group_count": receipt["approved_group_count"],
                "approved_source_rows": receipt["approved_source_rows"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
