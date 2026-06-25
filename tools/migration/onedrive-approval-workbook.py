#!/usr/bin/env python3
from __future__ import annotations

import argparse
import gzip
import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


DECISION_VALUES = "approve,needs_review,blocked,deferred,archive_only_override"
MATTER_TYPE_VALUES = "Criminal,Civil,Advisory,M&A"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--summary", required=True)
    parser.add_argument("--local-dir", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--receipt", required=True)
    return parser.parse_args()


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def read_ndjson_gz(path: Path) -> list[dict[str, Any]]:
    with gzip.open(path, "rt", encoding="utf-8") as handle:
        return [json.loads(line) for line in handle if line.strip()]


def join_list(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(map(str, value))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def evidence_summary(counts: dict[str, Any] | None) -> str:
    if not counts:
        return ""
    return ", ".join(f"{key}:{value}" for key, value in sorted(counts.items()) if value)


def priority_rank(priority: str) -> int:
    order = {
        "P0_conflict": 0,
        "P1_project_folder": 1,
        "P2_draft_code_ready": 2,
        "P3_missing_hint": 3,
        "P4_large_group": 4,
        "P5_standard_review": 5,
    }
    return order.get(priority, 9)


def priority_for(candidate: dict[str, Any]) -> str:
    warnings = set(candidate.get("code_warnings") or [])
    evidence_ref = candidate.get("evidence_ref") or {}
    project_like = int(evidence_ref.get("project_like_rows") or 0)
    object_count = int(candidate.get("object_count") or 0)
    if "conflicting_matter_type_hints" in warnings:
        return "P0_conflict"
    if project_like:
        return "P1_project_folder"
    if candidate.get("matter_code_candidate") and not warnings:
        return "P2_draft_code_ready"
    if "missing_client_or_matter_hint" in warnings:
        return "P3_missing_hint"
    if object_count >= 1000:
        return "P4_large_group"
    return "P5_standard_review"


def set_basic_style(sheet) -> None:
    normal_font = Font(name="Arial", size=10)
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = normal_font
            cell.alignment = wrap
            if cell.row > 1:
                cell.border = thin_border


def write_sheet(workbook: Workbook, name: str, rows: list[dict[str, Any]], freeze: str = "A2"):
    sheet = workbook.create_sheet(name)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    if not rows:
        sheet.append(["empty"])
    else:
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = freeze
        for index, header in enumerate(headers, start=1):
            preview_width = max((len(str(row.get(header) or "")) for row in rows[:200]), default=0)
            sheet.column_dimensions[get_column_letter(index)].width = max(
                12, min(max(len(str(header)), preview_width), 60) + 2
            )
    set_basic_style(sheet)
    return sheet


def build_rows(summary: dict[str, Any], local_dir: Path) -> dict[str, list[dict[str, Any]]]:
    reviewer = read_ndjson_gz(local_dir / "reviewer-packet.local.ndjson.gz")
    docs = read_ndjson_gz(local_dir / "document-census.local.ndjson.gz")
    candidates = read_ndjson_gz(local_dir / "candidate-hints.local.ndjson.gz")
    proposals = read_ndjson_gz(local_dir / "matter-code-proposals.local.ndjson.gz")
    target = read_ndjson_gz(local_dir / "target-resolution-plan.local.ndjson.gz")
    checklist = read_ndjson_gz(local_dir / "write-readiness-checklist.local.ndjson.gz")

    candidate_by_group = {row["group_id"]: row for row in candidates}
    proposal_by_group = {row["group_id"]: row for row in proposals}

    approval_rows: list[dict[str, Any]] = []
    for row in reviewer:
        group_id = row["group_id"]
        candidate = candidate_by_group.get(group_id, {})
        proposal = proposal_by_group.get(group_id, {})
        client_hint = candidate.get("client_hint") or {}
        matter_hint = candidate.get("matter_hint") or {}
        evidence_ref = candidate.get("evidence_ref") or {}
        priority = priority_for(candidate)
        approval_rows.append(
            {
                "review_priority": priority,
                "reviewer_decision": "needs_review",
                "approved_client_short_name": "",
                "approved_matter_type_english": "",
                "approved_matter_detail_type_korean": "",
                "approved_matter_code": "",
                "reviewer_notes": "",
                "group_id": group_id,
                "object_count": candidate.get("object_count") or row.get("object_count"),
                "project_like_rows": evidence_ref.get("project_like_rows", 0),
                "client_short_name_candidate": client_hint.get("client_short_name_candidate") or "",
                "client_hint_confidence": client_hint.get("confidence") or "",
                "matter_type_english_candidate": matter_hint.get("matter_type_english") or "",
                "matter_detail_type_korean_candidate": matter_hint.get("matter_detail_type_korean") or "",
                "matter_hint_confidence": matter_hint.get("confidence") or "",
                "matter_code_candidate": candidate.get("matter_code_candidate")
                or proposal.get("matter_code_candidate")
                or "",
                "code_warnings": join_list(candidate.get("code_warnings") or proposal.get("warnings") or []),
                "evidence_class_counts": evidence_summary(
                    evidence_ref.get("evidence_class_counts") or row.get("evidence_class_counts") or {}
                ),
                "raw_group_label_local_only": row.get("raw_group_label") or "",
                "required_action": DECISION_VALUES,
            }
        )
    approval_rows.sort(
        key=lambda item: (priority_rank(item["review_priority"]), -int(item["object_count"] or 0), item["group_id"])
    )

    priority_rows = [
        {
            "review_priority": row["review_priority"],
            "group_id": row["group_id"],
            "object_count": row["object_count"],
            "project_like_rows": row["project_like_rows"],
            "matter_code_candidate_present": bool(row["matter_code_candidate"]),
            "code_warnings": row["code_warnings"],
            "client_short_name_candidate": row["client_short_name_candidate"],
            "matter_type_english_candidate": row["matter_type_english_candidate"],
            "matter_detail_type_korean_candidate": row["matter_detail_type_korean_candidate"],
            "raw_group_label_local_only": row["raw_group_label_local_only"],
        }
        for row in approval_rows
    ]

    doc_rows = [
        {
            "group_id": row.get("group_id"),
            "sample_index": row.get("sample_index"),
            "leaf_name_local_only": row.get("leaf_name"),
            "title_hint_local_only": row.get("title_hint"),
            "extension": row.get("extension"),
            "size_bytes": row.get("size_bytes"),
            "evidence_classes": join_list(row.get("evidence_classes") or []),
            "content_read": row.get("content_read"),
            "ocr_excerpt_saved": row.get("ocr_excerpt_saved"),
            "screenshot_saved": row.get("screenshot_saved"),
        }
        for row in docs
    ]

    candidate_rows = []
    for row in candidates:
        client_hint = row.get("client_hint") or {}
        matter_hint = row.get("matter_hint") or {}
        evidence_ref = row.get("evidence_ref") or {}
        candidate_rows.append(
            {
                "group_id": row.get("group_id"),
                "review_state": row.get("review_state"),
                "object_count": row.get("object_count"),
                "client_short_name_candidate": client_hint.get("client_short_name_candidate"),
                "client_hint_confidence": client_hint.get("confidence"),
                "matter_type_english": matter_hint.get("matter_type_english"),
                "matter_detail_type_korean": matter_hint.get("matter_detail_type_korean"),
                "matter_hint_confidence": matter_hint.get("confidence"),
                "matter_code_candidate": row.get("matter_code_candidate"),
                "matter_code_format_valid": row.get("matter_code_format_valid"),
                "matter_code_length_valid": row.get("matter_code_length_valid"),
                "code_warnings": join_list(row.get("code_warnings") or []),
                "project_like_rows": evidence_ref.get("project_like_rows", 0),
                "evidence_class_counts": evidence_summary(evidence_ref.get("evidence_class_counts") or {}),
                "raw_group_label_local_only": row.get("raw_group_label"),
            }
        )

    proposal_rows = [
        {
            "group_id": row.get("group_id"),
            "review_state": row.get("review_state"),
            "client_short_name": row.get("client_short_name"),
            "matter_type_english": row.get("matter_type_english"),
            "matter_detail_type_korean": row.get("matter_detail_type_korean"),
            "matter_code_candidate": row.get("matter_code_candidate"),
            "format_valid": row.get("format_valid"),
            "length_valid": row.get("length_valid"),
            "unique_valid": row.get("unique_valid"),
            "warnings": join_list(row.get("warnings") or []),
        }
        for row in proposals
    ]

    target_rows = [
        {
            "group_id": row.get("group_id"),
            "review_state": row.get("review_state"),
            "target_resolution_state": row.get("target_resolution_state"),
            "client_resolution_state": row.get("client_resolution_state"),
            "matter_resolution_state": row.get("matter_resolution_state"),
            "matter_code_candidate_hash": row.get("matter_code_candidate_hash"),
            "blockers": join_list(row.get("blockers") or []),
        }
        for row in target
    ]

    checklist_rows = [
        {"gate": row.get("gate"), "status": row.get("status"), "required": row.get("required")}
        for row in checklist
    ]

    summary_rows = [
        {"metric": "original_rows", "value": summary.get("manifest_profile", {}).get("rows")},
        {"metric": "archive_only", "value": summary.get("final_readiness", {}).get("row_accounting", {}).get("archive_only")},
        {"metric": "needs_review", "value": summary.get("final_readiness", {}).get("row_accounting", {}).get("needs_review")},
        {"metric": "unsupported", "value": summary.get("final_readiness", {}).get("row_accounting", {}).get("unsupported")},
        {"metric": "reviewer_packet_groups", "value": len(reviewer)},
        {"metric": "representative_document_rows", "value": len(doc_rows)},
        {"metric": "candidate_hint_rows", "value": len(candidate_rows)},
        {"metric": "draft_matter_code_candidates", "value": sum(1 for row in proposal_rows if row.get("matter_code_candidate"))},
        {"metric": "approved_import_scope_rows", "value": 0},
    ]

    return {
        "Summary": summary_rows,
        "Approval": approval_rows,
        "Priority Queue": priority_rows,
        "Candidate Hints": candidate_rows,
        "Matter Code Proposals": proposal_rows,
        "Representative Docs": doc_rows,
        "Target Plan": target_rows,
        "Write Checklist": checklist_rows,
    }


def create_workbook(summary: dict[str, Any], rows_by_sheet: dict[str, list[dict[str, Any]]], output: Path) -> None:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme_rows = [
        ["AMIC OneDrive Bulk Approval Workbook", "local-only reviewer workbook"],
        ["Generated At", datetime.now(timezone.utc).isoformat()],
        ["Run ID", summary.get("run_id")],
        ["Original Rows", summary.get("manifest_profile", {}).get("rows")],
        ["Final Status", summary.get("final_readiness", {}).get("status")],
        ["Approval Boundary", "All generated candidates remain needs_review until reviewer approval."],
        ["Matter Code Format", "[client_short_name]/[matter_type_english]/[matter_detail_type_korean]"],
        ["Matter Type Values", MATTER_TYPE_VALUES],
        ["Decision Values", DECISION_VALUES],
        ["Security Boundary", "Workbook is local-only and may contain group labels and representative document names."],
    ]
    for row in readme_rows:
        readme.append(row)
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 110
    set_basic_style(readme)
    for cell in readme[1]:
        cell.font = Font(name="Arial", bold=True, size=13)

    approval_sheet = None
    for name, rows in rows_by_sheet.items():
        sheet = write_sheet(workbook, name, rows)
        if name == "Approval":
            approval_sheet = sheet

    if approval_sheet is not None:
        headers = [cell.value for cell in approval_sheet[1]]
        if "reviewer_decision" in headers:
            column_index = headers.index("reviewer_decision") + 1
            column_letter = get_column_letter(column_index)
            validation = DataValidation(type="list", formula1=f'"{DECISION_VALUES}"', allow_blank=False)
            approval_sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}{approval_sheet.max_row}")

        if "approved_matter_type_english" in headers:
            column_index = headers.index("approved_matter_type_english") + 1
            column_letter = get_column_letter(column_index)
            validation = DataValidation(type="list", formula1=f'"{MATTER_TYPE_VALUES}"', allow_blank=True)
            approval_sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}{approval_sheet.max_row}")

        input_columns = {
            "reviewer_decision",
            "approved_client_short_name",
            "approved_matter_type_english",
            "approved_matter_detail_type_korean",
            "approved_matter_code",
            "reviewer_notes",
        }
        fill = PatternFill("solid", fgColor="FFF2CC")
        for index, header in enumerate(headers, start=1):
            if header in input_columns:
                for cell in approval_sheet.iter_cols(min_col=index, max_col=index, min_row=2, max_row=approval_sheet.max_row):
                    for target in cell:
                        target.fill = fill

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = True
    workbook.save(output)
    os.chmod(output, 0o600)


def validate_workbook(output: Path, receipt: Path, rows_by_sheet: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    workbook = load_workbook(output, read_only=True, data_only=False)
    expected = ["README", *rows_by_sheet.keys()]
    validation = {
        "artifact": "approval_workbook_validation_sanitized",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "workbook_ref": output.name,
        "mode": oct(os.stat(output).st_mode & 0o777)[2:],
        "sheet_names": workbook.sheetnames,
        "expected_sheets_present": all(sheet in workbook.sheetnames for sheet in expected),
        "row_counts": {name: workbook[name].max_row - 1 for name in rows_by_sheet},
        "approved_import_scope_rows": 0,
        "security_boundary": "local-only workbook; may contain raw group labels and representative document names; do not commit",
        "repo_safe": False,
    }
    validation["checks"] = {
        "mode_0600": validation["mode"] == "600",
        "expected_sheets_present": validation["expected_sheets_present"],
        **{
            f"{name.lower().replace(' ', '_')}_rows_match": validation["row_counts"][name] == len(rows)
            for name, rows in rows_by_sheet.items()
        },
    }
    validation["all_checks_passed"] = all(validation["checks"].values())
    receipt.write_text(json.dumps(validation, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.chmod(receipt, 0o644)
    return validation


def main() -> None:
    args = parse_args()
    summary_path = Path(args.summary)
    local_dir = Path(args.local_dir)
    output = Path(args.output)
    receipt = Path(args.receipt)
    output.parent.mkdir(parents=True, exist_ok=True)
    receipt.parent.mkdir(parents=True, exist_ok=True)

    summary = read_json(summary_path)
    rows_by_sheet = build_rows(summary, local_dir)
    create_workbook(summary, rows_by_sheet, output)
    validation = validate_workbook(output, receipt, rows_by_sheet)
    print(
        json.dumps(
            {
                "workbook": str(output),
                "receipt": str(receipt),
                "all_checks_passed": validation["all_checks_passed"],
                "row_counts": validation["row_counts"],
                "mode": validation["mode"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
