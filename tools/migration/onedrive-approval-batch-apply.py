#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DECISIONS = {"approve", "needs_review", "blocked", "deferred", "archive_only_override"}
APPLY_COLUMNS = [
    "reviewer_decision",
    "approved_client_short_name",
    "approved_matter_type_english",
    "approved_matter_detail_type_korean",
    "approved_matter_code",
    "reviewer_notes",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-workbook", required=True)
    parser.add_argument("--batch-workbook", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--receipt", required=True)
    return parser.parse_args()


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def header_map(sheet) -> dict[str, int]:
    return {clean(cell.value): index for index, cell in enumerate(sheet[1], start=1)}


def read_batch_decisions(batch_workbook: Path) -> tuple[dict[str, dict[str, str]], dict[str, Any]]:
    workbook = load_workbook(batch_workbook, read_only=True, data_only=False)
    if "Review Batch" not in workbook.sheetnames:
        return {}, {"blockers": ["missing_review_batch_sheet"]}
    sheet = workbook["Review Batch"]
    headers = header_map(sheet)
    required = {"group_id", "batch_decision", "approved_client_short_name", "approved_matter_type_english", "approved_matter_detail_type_korean", "approved_matter_code", "reviewer_notes"}
    missing = sorted(required - set(headers))
    if missing:
        return {}, {"blockers": [f"missing_batch_column:{column}" for column in missing]}
    decisions: dict[str, dict[str, str]] = {}
    counts: Counter[str] = Counter()
    invalid_rows = []
    for row_index in range(2, sheet.max_row + 1):
        group_id = clean(sheet.cell(row_index, headers["group_id"]).value)
        decision = clean(sheet.cell(row_index, headers["batch_decision"]).value) or "needs_review"
        counts[decision] += 1
        if decision not in DECISIONS:
            invalid_rows.append(row_index)
            continue
        if not group_id:
            invalid_rows.append(row_index)
            continue
        decisions[group_id] = {
            "reviewer_decision": decision,
            "approved_client_short_name": clean(sheet.cell(row_index, headers["approved_client_short_name"]).value),
            "approved_matter_type_english": clean(sheet.cell(row_index, headers["approved_matter_type_english"]).value),
            "approved_matter_detail_type_korean": clean(sheet.cell(row_index, headers["approved_matter_detail_type_korean"]).value),
            "approved_matter_code": clean(sheet.cell(row_index, headers["approved_matter_code"]).value),
            "reviewer_notes": clean(sheet.cell(row_index, headers["reviewer_notes"]).value),
        }
    return decisions, {"blockers": [], "decision_counts": dict(sorted(counts.items())), "invalid_row_count": len(invalid_rows)}


def copy_batch_values(base_sheet, headers: dict[str, int], row_index: int, values: dict[str, str]) -> None:
    for column in APPLY_COLUMNS:
        if column in headers:
            base_sheet.cell(row_index, headers[column]).value = values.get(column, "")


def main() -> None:
    args = parse_args()
    base_workbook = Path(args.base_workbook)
    batch_workbook = Path(args.batch_workbook)
    output = Path(args.output)
    receipt = Path(args.receipt)
    output.parent.mkdir(parents=True, exist_ok=True)
    receipt.parent.mkdir(parents=True, exist_ok=True)

    batch_decisions, batch_meta = read_batch_decisions(batch_workbook)
    workbook = load_workbook(base_workbook)
    blockers = list(batch_meta.get("blockers") or [])
    if "Approval" not in workbook.sheetnames:
        blockers.append("missing_approval_sheet")

    applied_rows = 0
    missing_group_ids = []
    base_decision_counts: Counter[str] = Counter()
    output_decision_counts: Counter[str] = Counter()
    if not blockers:
        sheet = workbook["Approval"]
        headers = header_map(sheet)
        required = {"group_id", *APPLY_COLUMNS}
        missing = sorted(required - set(headers))
        if missing:
            blockers.extend(f"missing_approval_column:{column}" for column in missing)
        else:
            seen_groups = set()
            for row_index in range(2, sheet.max_row + 1):
                group_id = clean(sheet.cell(row_index, headers["group_id"]).value)
                base_decision_counts[clean(sheet.cell(row_index, headers["reviewer_decision"]).value) or "blank"] += 1
                if group_id in batch_decisions:
                    copy_batch_values(sheet, headers, row_index, batch_decisions[group_id])
                    applied_rows += 1
                    seen_groups.add(group_id)
                output_decision_counts[clean(sheet.cell(row_index, headers["reviewer_decision"]).value) or "blank"] += 1
            missing_group_ids = sorted(set(batch_decisions) - seen_groups)
            if missing_group_ids:
                blockers.append("batch_group_ids_missing_from_base_workbook")

    workbook.save(output)
    os.chmod(output, 0o600)

    receipt_payload = {
        "artifact": "approval_batch_apply_sanitized",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "base_workbook_ref": base_workbook.name,
        "batch_workbook_ref": batch_workbook.name,
        "output_workbook_ref": output.name,
        "mode": oct(os.stat(output).st_mode & 0o777)[2:],
        "status": "applied" if not blockers else "blocked",
        "batch_decision_counts": batch_meta.get("decision_counts", {}),
        "base_decision_counts_before": dict(sorted(base_decision_counts.items())),
        "output_decision_counts_after": dict(sorted(output_decision_counts.items())),
        "batch_rows_matched": applied_rows,
        "missing_batch_group_count": len(missing_group_ids),
        "blockers": blockers,
        "security_boundary": "local-only workbook; may contain raw group labels; do not commit",
        "repo_safe": False,
    }
    receipt_payload["checks"] = {
        "output_mode_0600": receipt_payload["mode"] == "600",
        "no_blockers": not blockers,
        "all_batch_rows_matched": len(missing_group_ids) == 0,
    }
    receipt_payload["all_checks_passed"] = all(receipt_payload["checks"].values())
    receipt.write_text(json.dumps(receipt_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.chmod(receipt, 0o644)
    print(
        json.dumps(
            {
                "status": receipt_payload["status"],
                "all_checks_passed": receipt_payload["all_checks_passed"],
                "batch_rows_matched": applied_rows,
                "output_decision_counts_after": receipt_payload["output_decision_counts_after"],
                "output": str(output),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
